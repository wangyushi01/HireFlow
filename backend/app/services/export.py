import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import Candidate

STATUS_LABELS = {
    "resume_received": "已投递",
    "screening": "筛选中",
    "interview": "面试中",
    "offer": "Offer",
    "hired": "已入职",
    "rejected": "已拒绝",
    "talent_pool": "人才库",
}

HEADERS = [
    "ID", "姓名", "手机", "邮箱", "应聘岗位", "状态", "来源",
    "学历", "学校", "工作年限", "当前公司", "期望薪资",
    "到岗时间", "技能", "AI 评分", "AI 推荐", "AI 评价",
    "投递时间", "简历文件",
]

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_candidates_excel(db: Session, status: str | None = None,
                            job_id: int | None = None, source: str | None = None) -> io.BytesIO:
    q = db.query(Candidate)
    if status:
        q = q.filter(Candidate.status == status)
    if job_id:
        q = q.filter(Candidate.job_id == job_id)
    if source:
        q = q.filter(Candidate.source == source)
    candidates = q.order_by(Candidate.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "候选人列表"

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, c in enumerate(candidates, 2):
        job_title = c.job.title if c.job else ""
        values = [
            c.id, c.name, c.phone or "", c.email or "", job_title,
            STATUS_LABELS.get(c.status, c.status), c.source or "",
            c.highest_degree or "", c.school or "",
            c.years_of_experience or "", c.current_company or "",
            c.expected_salary or "", c.availability or "",
            ", ".join(c.get_skills()),
            c.ai_score if c.ai_score is not None else "",
            c.ai_recommendation or "", c.ai_summary or "",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
            c.resume_filename or "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["N"].width = 30
    ws.column_dimensions["Q"].width = 25
    ws.column_dimensions["R"].width = 14

    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(candidates) + 1}"
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
